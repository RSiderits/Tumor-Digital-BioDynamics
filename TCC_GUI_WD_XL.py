
# Title: Tumor Cell DNA Section Estimation

# Author Richard Siderits 

# Version 1.2 (7/26/22)
# Version 1.1 (7/17/22)    


 # Notes: ------------------------------------------------------------------ @
    
    # Human diploid cell in G1 has about 6 picograms of DNA
    
    # Nuclear DNA Content Varies with Cell Size, across Human Cell Types
          
    # Non-tumor cellularity - epithelial, mesenchymal, immune, endothelial, 
    #fribrblast)
    
    # TIS = Tumor Induced Stroma (Tumor microenvironment)
    
    # TCF = Tumor Cell Fields

    # IBTC = In Between Tumor Cells

    # PNT = Parenchyma Non-Tumor

    # TAS = Tumor Associated Stroma (reparative)
    
    
 # Assumptions: ------------------------------------------------------------ @
     
    # The tumor is roughly spherical.
    # Sections at equitorial plane of section (standard grossing practice) 
    # Picograms DNA are from Enhanced for tumor samples (stroma rmoved)
    # Count necrosis as stromal for estimates of non-tumor
    # Average cell has about 6 pg DNA
    # Non-tumor cellularity is epithelial, mesenchymal, immune, endothelial, 
    # fribrblast)
    # "Enhanced for tumor" follows microdissection, non-tumor stroma is 
    # excised from block
    # Extraction efficiency and proportion of Tumor Induced Stroma (TIS) vary.
    # Necrosis in Tumor Cell Field (TCF), should be counted as non-tumor.
    
 # Instructions: ----------------------------------------------------------- @
        
    #     
    #
    #     1) Review tissue in H&E or Hematoxolyn stained sections.
    #     2) Estimate:
    #         - Average diameter of the lesion or TCF if using a core
    #         - Average size of the tumor cells in um, measure up o 20 cells.
    #         - Percent TIS
    #         - Cellularity of TIS, include areas of necrosis
    #     3) If you are using a core, then use the core diameter as leasion 
    #        size
    #     4) Enter variables into the respective fields then click "calculate"

    # The final field will provide an estmated number of 10um thick secton to 
    # assure 100ng of Tumor DNA 
    
    # A report in MS Word format can be genereated in the same directory.  
    # 
    # Click on Excel button to send reuslts to an Excel file in the working 
    # directory
    #
    #

 # -------------------------- Import packages ------------------------------ @

import datetime
import docxtpl 
from docxtpl import DocxTemplate
import PySimpleGUI as sg   
import math

from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection 
from openpyxl.styles import Font, colors, Alignment

 # -------------------------- Create Workbook ----------------------------- @

wb = load_workbook(filename = 'pyXL.xlsx')

ws = wb.active
ws.title = "Log Data"
ws.sheet_properties.tabColor = "A3D03B"
ws1 = wb["Log Data"]

 # -------------------------- Border -------------------------------------- @

thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

ws1["A1"]="ID"
ws1["B1"]="CUTS"
ws1["A1"].border=thick_border
ws1["B1"].border=thick_border

align=Alignment(horizontal="center")

ws["A1"].alignment = Alignment(horizontal="center")
ws["B1"].alignment = Alignment(horizontal="center")

ws['A1'].fill = PatternFill(start_color="FFC7CE", fill_type = "solid")
ws['B1'].fill = PatternFill(start_color="FFC7CE", fill_type = "solid")  
                                            
 # --------------------------- Date ---------------------------------- @

today=datetime.date.today()

document_path = Path(__file__).parent / "RPT1TMPLT.docx"

doc = DocxTemplate("RPT1TMPLT.docx") 

 # --------------------------- Functions ----------------------------- @

def check():
          
      if event =="-ID-" :
      
          if values["-TS-"]== "":
          
           window["-TS-"].update("1") 
              
           window["-CS-"].update("10")            
          
           window["-PS-"].update("10") 
          
           window["-NT-"].update("10") 
          
           window["-EE-"].update("80")    
         
          return 

  
 # -------------------------- Variables ----------------------------------- @


tumor_size=0 # in cm
tumor_size_um=0 # Calculated
tumor_volume=0 # Calculated
tumor_radius=0 # Calculated
percent_stroma=0 # Input as a percent.
percent_nontumor_cellularity=0 # Input as a percent
PT=0 # Picograms of tumor genomic material
comp_number_cells=0

cell_radius=0 # Calculated
cell_volume=0 # Calculated
cell_size=0  # in microns
CN=0  # Compensated nummber of tumor cells
NT=0  # Non tumor, stroma TIS and Necrosis
number_cells=0
VC=0   # Volume cylinder of sections, DNA retreival, 10, 10 micron sections
CC =0 # Cells in Cylinder
PDC=0 # Picogram DNA in cylinder of sections for sample
pi=math.pi
strmlDNA=0 # Non-Tumoor DNA
PDC=0 # Picograms DNA in a 100 um high cylinder
EEC=0
sgDNA=0
v1,v2,v3,v4,v5=0,0,0,0,0
EECNE=0 # Non-Enhanced for tumor with Extraction efficiency compensated
EECNENG=0 # Convert to Nanograms
sg.theme("Topanga")


 # ------------------------ Layout ---------------------------------------- @

layout=[
        [sg.Image(filename=("DNA2.png"), size=(200,200), pad=(20,10)), 
         
         sg.Text("Sample ID:", border_width=3, background_color="black", 
                 text_color="yellow"), 
        
         sg.Input("1", size=(15,1), border_width=7, 
         text_color="yellow", enable_events=True, focus=True, key="-ID-", 
         background_color="black")],
       
        # Get input for parameters: --------------------------------------- @
        
        [sg.Text("Press [Tab] key to advance:", text_color="gold")],
        [sg.Text()],                    
        [sg.Input(default_text="1", key="-TS-", size=(5,1), 
                  text_color="green1", 
                  enable_events=True,tooltip=" Enter size of lesion in cm" ),    
        sg.Text("Tumor size in cm (1-10 cm)", size=(25,1))],  
        
        [sg.Input(default_text="25",key="-CS-", size=(5,1), 
                  text_color="green1", 
                  enable_events=True),
        sg.Text("Average tumor cell size in microns (1-100 um)",
                size=(35,1))], 
       
        [sg.Input(default_text="10",key="-PS-", size=(5,1), 
                  text_color="green1", 
                  enable_events=True), 
        sg.Text("Percent stroma (1-99%)", size=(20,1))],  
  
        [sg.Input(default_text="10", key="-NT-", size=(5,1), 
                  text_color="green1", 
                  enable_events=True),
        sg.Text("Percent stromal cellularity (1-99%)", size=(25,1))],         
   
        [sg.Input(default_text="85",key="-EE-", size=(5,1), 
                  text_color="green1", 
                  enable_events=True),
         
        sg.Text("Expected Extraction efficiency (1-99%)", size=(30,1))],        
      
       
        [sg.Text("--------------------------------------------------------")],
        
       
        
        # ----------- Clear, Write Word report, send to XL and quit ------- @
        
        
        [sg.Button("Calculate", key="-Calc-", pad=(10,20), size=20, 
                   font="Arial,12"), 
         
         sg.Button("Clear", font="Arial,12", key="-Clear-", pad=(10,25)),  
         
         sg.Exit(font="Arial,12", pad=(10,25)),
         
         sg.Button("Write Report", font="Arial,12", key="-RPT-", size=12, 
                   pad=(10,10)),
         
         sg.Button("Send Excel", font="Arial,12", key="S2XL", size=11, 
                   pad=(10,10))
        ],
        
              
        [sg.Text("--------------------------------------------------------")],
       
        # Output fileds for calculations, in order EN,CN,PT --------------- @
                        
        [sg.Text("Estimated number of tumor cells:"), 
         sg.Text("___", size=(30,1), key="-EN-", text_color="gold")], 
     
        [sg.Text("Stroma Compensated number of tumor cells:"), 
         sg.Text("___", size=(30,1), key="-CN-" , text_color="gold")], 
        
        [sg.Text("Estimated total picograms tumor DNA:"), 
         sg.Text("0.0", size=(10,1), key="-PT-", text_color="gold")], 
        
        [sg.Text("--------------------------------------------------------")],
        
        [sg.Text("Estimated DNA if Enhcanced for Tumor, in a 100um thick\
 section pg DNA:"), 
         sg.Text("0.0", size=(10,1), key="-PDC-", text_color="gold")], 
        
        [sg.Text("STROMAL DNA for 100um thick sample, not enhanced for tumor\
 pg DNA:"), 
         sg.Text("0.0", size=(10,1), key="-STDNA-", text_color="gold")],

        [sg.Text("Compensated for Extraction Efficiency, All cells in sample,\
 pg DNA:"), 
         sg.Text("0.0", size=(10,1), key="-EEC-", text_color="gold")],

        [sg.Text("Sample Not Enhanced for Tumor, but Extraction Efficiency\
 compensated, pg DNA:"), 
         sg.Text("0.0", size=(10,1), key="-EECNE-", text_color="gold")],
        
        [sg.Text("Sample Not Enhanced for Tumor, Efficiency compensated,\
 convert to ng DNA:"), 
         sg.Text("0.0", size=(10,1), key="-EECNENG-", text_color="gold")],
        
        [sg.Text("# of 10um thick sections, enhanced, to assure 100 ng gDNA\
 for sequencing:"), 
         sg.Text("0.0", size=(10,1), key="-sgDNA-", text_color="gold")],

        [sg.Text("--------------------------------------------------------")],
        [sg.Text("")],
        
       ]        


 # ------------------------ Create the window ----------------------------- @
 

window = sg.Window("Estimate Tumor Cell Count and amount tumor DNA", layout)


 # ------------------------ Process events -------------------------------- @

while True:
    
    event, values = window.read()
    
    
    if event == sg.WIN_CLOSED or event == "Exit":
      break
      
    
    if event == "-TS-":
         
         window['-TS-'].update(background_color="navy")
         
    if event == "-CS-":
          
        window['-CS-'].update(background_color="navy")
    
    if event == "-PS-":
        window['-PS-'].update(background_color="navy")
    
    if event == "-NT-":
                  
       window['-NT-'].update(background_color="navy")
       
    if event == "-EE-":
                     
       window['-EE-'].update(background_color="navy")
    
       
    if event == "-Calc-":
              
        v1=values["-TS-"] # Tumor size in cm
        if values["-TS-"]=="":
            v1=int(1)
            window["-TS-"].update(v1)
            
        v2=values["-PS-"] # Percent stroma (TIS)
        if values["-PS-"]=="":
            v2=int(25)
            window["-PS-"].update(v2)
            
        v3=values["-CS-"] # Average tumor cell size in microns
        if values["-CS-"]=="":
            v3=int(25)
            window["-CS-"].update(v3)
                       
        v4=values["-NT-"] # Percent non-tumor cellularity 
        if values["-NT-"]=="":
            v4=int(10)
            window["-NT-"].update(v3)
        
        v5=values["-EE-"]
        if values["-EE-"]=="":
            v5=int(80)
            window["-EE-"].update(v3)
       

       #check()        
              
        #window.refresh()
       
     
 # ----------- Do the calculations with the input values ------------------ @             
  
    
     
 # Tumor size in microns:
        tumor_size_um = float(v1)*1000
        
 # Tumor radius in microns:  
        tumor_radius = float(tumor_size_um)*0.5
        
 # Tumor volume in cubic microns V=4/3 πr**3
        tumor_volume = int((4/3)*pi*tumor_radius**3)
        
                                 # ---------------------------------------- @
 # Cell radius of tumor cell in microns:  
        cell_radius = float(v3)*0.5
        
 # Cell volume tumor cells in cubic microns:  
        cell_volume = int((4/3)*pi*cell_radius**3)
        
                                 # ---------------------------------------- @
                
 # Number of cells in tumor for pure tumor cells: 
        number_cells = int(tumor_volume/cell_volume)
    
 # Number of cells in tumor, Stroma Compensated :
        perstroma=(100-int(v2))/100
        comp_number_cells = int(number_cells*perstroma)    
        
 # Estimated PicoGrams of tumor DNA in entire Tumor (TCF).
        PT=int((comp_number_cells*6)) # Picograms DNA from stroma compensated.
        
        # Picograms DNA in a cylinder of 100 um high (10, 10 um sections)
        #   h=100 um
        #   r= tumor radius (tumor_radius) in um
        #   Pi from math module
        #   VC=pi r**2 h  VC is volume of a cylinder.
        
 # Cylinder Volume for 100um high 10, 10um sections:
        VC=int(pi*(tumor_radius**2)*100) # cubic microns
        
 # Number of cells in the sample cylinder, NOT compensated for stroma
        CC=int(VC/cell_volume)
        
 # Stromal DNA Non-Tumor, compensated for %Stroma and Stromal cellularity
        strmlDNA=int(float(VC/cell_volume*(perstroma/100*float(v4)*6)))
       
 # Picograms of DNA in cylinder, if enhanced for tumor
        PDC=int(CC*6)
        
 # Compensated for Extraction efficinecy, ENHANCED for tumor:
     
        EEC= PDC*(int(v5)/100) 
        
 # Compensated for Extraction efficinecy, NOT ENHANCED for tumor:
     
        EECNE=int((PDC-strmlDNA)*(int(v5)/100))
        
        EECNENG=EECNE/1000
        
 # Number 10um thick sections of enhanced for tumor sample, assure 100ng gDNA
 
        sgDNA=int(round((100/EECNENG)*10,2)) 
        
  
       
 # ----------  SEND DATA TO Window ---------------------------------------- @   
        
        window["-EN-"].update('{:,}'.format(number_cells))  
        window["-CN-"].update('{:,}'.format(comp_number_cells))  
        window["-PT-"].update('{:,}'.format(PT))  
        window["-PDC-"].update('{:,}'.format(PDC))
        window["-STDNA-"].update('{:,}'.format(strmlDNA)) 
        window["-EEC-"].update('{:,}'.format(EEC)) 
        window["-EECNE-"].update('{:,}'.format(EECNE)) 
        window["-EECNENG-"].update('{:,}'.format(EECNENG)) 
        window["-sgDNA-"].update('{:,}'.format(sgDNA)) 
   
    
 # ----------  CLEAR Fields ----------------------------------------------- @      
     
    if event == "-Clear-":
        
        window["-TS-"].update("")
        window["-PS-"].update("")
        window["-CS-"].update("")
        window["-CN-"].update("")
        window["-EN-"].update("")
        window["-PT-"].update("")
        window["-PDC-"].update("")
        window["-NT-"].update("")
        window["-EE-"].update("")
        window["-EEC-"].update("")
        window["-STDNA-"].update("")
        window["-EECNE-"].update("")
        window["-EECNENG-"].update("")
        window["-sgDNA-"].update("")
        window["-ID-"].update("")
        window['-TS-'].update(background_color="black")
        window['-CS-'].update(background_color="black")
        window['-PS-'].update(background_color="black")
        window['-NT-'].update(background_color="black")
        window['-EE-'].update(background_color="black")
        
        # Event Values are in a Dictionary not a list.
        
 # ---------  Write MS Word Report ---------------------------------------- @         
  
    if event == "-RPT-": 
        
        SID=values["-ID-"]
                
        context= {
            "ID": SID,
            "DTR": today,
            "TSize": v1,
            "TSizeM": v3,
            "PStroma":v2,
            "PSC": v4,
            "EE":v5,
            "ENTC": number_cells,
            "SCTC": comp_number_cells,
            "TPDNA": PT,
            "CDNA": PDC,
            "CSDNA":strmlDNA,
            "CCEE": EEC,
            "CNEEE": EECNE,
            "SING": EECNENG,
            "sgDNA": sgDNA,
                 }
                          
        doc.render(context)
                
        doc.save(Path(__file__).parent / f"A_Sample__{SID}__.docx") 
                       
        sg.theme("DarkBlue1")
        
        sg.popup("Notification:",  f"File saved to {document_path}", 
                 text_color="yellow", font=("ariel", 14), auto_close=True, 
                 auto_close_duration=3 )

 # ---------------------------- EXCEL-------------------------------------- @  

    if event == "S2XL":
        
        
        v1=values["-ID-"]
        v2=sgDNA
        
            
        ws.append([v1,v2])
        
        wb.save("pyXL.xlsx") 
        
        sg.popup("Notification:", "Saved to pyXL.xlsx working directory", 
                 text_color="yellow", font=("ariel", 14), auto_close=True, 
                 auto_close_duration=2 )
        
           
        window["-ID-"].set_focus()
        
wb.save("pyXL.xlsx")   

window.close()

    
    

